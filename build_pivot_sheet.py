#!/usr/bin/env python3
"""
build_pivot_sheet.py
====================

Rebuilds "Sheet1" of the social-feed workbook as a REAL, refreshable Excel
PivotTable -- not a static copy of the numbers.

Pivot definition being reproduced
---------------------------------
    Source      : sheet "social_feed_insights (41)", range A1:T<last>
    Rows        : page_name  ->  hashtag        (compact / outline layout)
    Values      : Count of message
                  Sum of approximate_earnings
                  Average of video_views
    Layout      : subtotals on top of each page_name group, grand total row,
                  style PivotStyleLight16, comma [0] number format
    Anchor      : A3

What the script writes
----------------------
The output .xlsx contains

    1. a data sheet holding the source rows as static values, and
    2. "Sheet1" carrying a genuine PivotTable part
       (xl/pivotTables/pivotTable1.xml + xl/pivotCache/*) that is linked to
       that data sheet.  You can click it in Excel, open the Field List,
       drag fields around, and hit Refresh.

The visible grid is also written into the cells so the sheet renders
correctly even before the first refresh -- exactly what Excel itself does.

Usage
-----
    python build_pivot_sheet.py INPUT.(csv|xlsx) [-o OUTPUT.xlsx]
                                [--source-sheet "social_feed_insights (41)"]
                                [--pivot-sheet Sheet1]

or from Python:

    from build_pivot_sheet import build
    build("facebook_export_csvs/social_feed_insights_....csv")

Requires: openpyxl >= 3.1
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import re
import sys
from collections import OrderedDict
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.pivot.cache import (
    CacheDefinition,
    CacheField,
    CacheSource,
    SharedItems,
    WorksheetSource,
)
from openpyxl.pivot.fields import Boolean, DateTimeField, Index, Missing, Number, Text
from openpyxl.pivot.record import Record, RecordList
from openpyxl.pivot.table import (
    DataField,
    FieldItem,
    Location,
    PivotField,
    PivotTableStyle,
    RowColField,
    RowColItem,
    TableDefinition,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Pivot configuration -- change these four constants to re-point the pivot
# --------------------------------------------------------------------------

ROW_FIELDS = ["page_name", "hashtag"]

# (source column, aggregation, caption)  -- order == column order B, C, D
VALUE_FIELDS = [
    ("message", "count", "Count of message"),
    ("approximate_earnings", "sum", "Sum of approximate_earnings"),
    ("video_views", "average", "Average of video_views"),
]

PIVOT_ANCHOR_ROW = 3          # pivot header lands on row 3, as in the original
PIVOT_ANCHOR_COL = 1          # column A
PIVOT_STYLE = "PivotStyleLight16"

# "Comma [0]" -- exact zeros render as "-", 0.4 renders as "0", matching the
# original sheet.
COMMA_0 = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

BLANK_CAPTION = "(blank)"


# --------------------------------------------------------------------------
# Reading the source
# --------------------------------------------------------------------------

_INT_RE = re.compile(r"^[+-]?\d+$")
_DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_DATETIME_RE = re.compile(
    r"^(\d{4})-(\d{2})-(\d{2})[T ](\d{2}):(\d{2})(?::(\d{2}))?"
)


def _coerce(text: str) -> Any:
    """Turn a CSV string into the int / float / datetime / str Excel would infer."""
    s = text.strip()
    if not s:
        return None
    m = _DATETIME_RE.match(s)
    if m:
        y, mo, d, hh, mm, ss = m.groups()
        return _dt.datetime(int(y), int(mo), int(d), int(hh), int(mm), int(ss or 0))
    m = _DATE_RE.match(s)
    if m:
        return _dt.datetime(*(int(g) for g in m.groups()))
    if _INT_RE.match(s):
        n = int(s)
        # Beyond 2**53 Excel loses precision anyway; keep those as text.
        return n if abs(n) < 2 ** 53 else s
    try:
        f = float(s)
    except ValueError:
        return s
    # "nan"/"inf" parse as floats but are not what a spreadsheet means.
    return f if f == f and abs(f) != float("inf") else s


def _safe_sheet_name(name: str) -> str:
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    return name[:31] or "data"


def read_csv_source(path: str, sheet_name: str | None):
    """Read the exporter's social_feed_insights_*.csv straight into memory."""
    import csv as _csv

    delim = "\t" if path.lower().endswith((".tsv", ".tab")) else ","
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = _csv.reader(fh, delimiter=delim)
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            raise SystemExit("csv is empty")
        while headers and not headers[-1]:
            headers.pop()
        width = len(headers)

        rows: list[list[Any]] = []
        for raw in reader:
            if not any(c.strip() for c in raw):
                continue
            vals = list(raw[:width]) + [""] * max(0, width - len(raw))
            rows.append([_coerce(v) for v in vals])

    if not rows:
        raise SystemExit("csv has no data rows")

    if sheet_name is None:
        stem = os.path.splitext(os.path.basename(path))[0]
        sheet_name = _safe_sheet_name(stem)
    return sheet_name, headers, rows


def read_source(path: str, sheet_name: str | None):
    """Return (sheet_name, headers, rows) from either a .csv or a .xlsx.

    For xlsx, data_only=True is essential: `hashtag` is a formula column in the
    original workbook, so we need Excel's cached results, not the formula text.
    """
    if path.lower().endswith((".csv", ".tsv", ".tab")):
        return read_csv_source(path, sheet_name)

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name is None:
        sheet_name = next(
            (n for n in wb.sheetnames if n.lower().startswith("social_feed")),
            wb.sheetnames[0],
        )
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    it = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(it)
    except StopIteration:
        raise SystemExit("source sheet is empty")

    # Trim trailing all-empty header columns.
    headers = [("" if h is None else str(h)) for h in raw_headers]
    while headers and not headers[-1]:
        headers.pop()
    width = len(headers)

    rows: list[list[Any]] = []
    for r in it:
        vals = list(r[:width]) + [None] * max(0, width - len(r))
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue  # skip fully blank rows
        rows.append([_norm(v) for v in vals])

    wb.close()
    if not rows:
        raise SystemExit("source sheet has no data rows")
    return sheet_name, headers, rows


def _norm(v: Any) -> Any:
    """Empty strings behave as blanks, matching Excel's pivot cache."""
    if isinstance(v, str) and not v.strip():
        return None
    if isinstance(v, _dt.datetime):
        return v
    if isinstance(v, _dt.date):
        return _dt.datetime(v.year, v.month, v.day)
    return v


# --------------------------------------------------------------------------
# Aggregation (for the rendered cell values; Excel recomputes on refresh)
# --------------------------------------------------------------------------

def item_key(v):
    """The identity Excel uses for a pivot item.

    PivotTable items are matched case-INSENSITIVELY: "#HindiStory" and
    "#hindistory" are one item, not two. Blanks are their own item.
    """
    return None if v is None else str(v).casefold()


def sort_key(k):
    """Excel's default item order: ascending, blanks last."""
    return (1, "") if k is None else (0, k)


def build_item_index(rows, col):
    """Return (labels, position) for one axis column.

    `labels` are the captions in display order -- for a case-collision group
    Excel shows the casing it saw FIRST in the source, so that is what we keep.
    `position` maps item_key -> index into `labels`, and that index is what
    both the cache records and the pivot's rowItems refer to. Deriving both
    from one place is what keeps them consistent; the previous version
    re-derived the index with list.index() per row, which mis-ordered items
    whenever two values differed only by case.
    """
    first = {}
    for r in rows:
        k = item_key(r[col])
        if k not in first:
            first[k] = r[col]
    keys = sorted(first, key=sort_key)
    return [first[k] for k in keys], {k: i for i, k in enumerate(keys)}


def aggregate(headers, rows, item_index):
    """Build the row tree and the numbers shown in each cell."""
    ri = [headers.index(f) for f in ROW_FIELDS]
    vi = [headers.index(c) for c, _, _ in VALUE_FIELDS]

    outer_labels, outer_pos = item_index[ri[0]]
    inner_labels, inner_pos = item_index[ri[1]]

    # Keyed by item INDEX, so grid order and pivot item order cannot drift.
    tree: dict[int, dict[int, list]] = {}
    for r in rows:
        o = outer_pos[item_key(r[ri[0]])]
        i = inner_pos[item_key(r[ri[1]])]
        tree.setdefault(o, {}).setdefault(i, []).append(r)

    def measure(bucket, col_idx, how):
        vals = [b[col_idx] for b in bucket]
        if how == "count":                       # COUNTA -- non-blank
            return sum(1 for v in vals if v is not None)
        nums = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if how == "sum":
            return sum(nums)
        if how == "average":
            return (sum(nums) / len(nums)) if nums else None
        raise ValueError(how)

    def measures(bucket):
        return [measure(bucket, vi[k], VALUE_FIELDS[k][1]) for k in range(len(VALUE_FIELDS))]

    # (indent, label, [values], is_subtotal, is_grand_total, item_index)
    grid = []
    for o in sorted(tree):
        inners = tree[o]
        flat = [r for b in inners.values() for r in b]
        grid.append((0, outer_labels[o], measures(flat), True, False, o))
        for i in sorted(inners):
            grid.append((1, inner_labels[i], measures(inners[i]), False, False, i))
    grid.append((0, "Grand Total", measures(rows), True, True, None))
    return tree, grid


# --------------------------------------------------------------------------
# Pivot cache
# --------------------------------------------------------------------------

def build_cache(headers, rows, source_sheet, item_index):
    """Create the pivotCacheDefinition + pivotCacheRecords.

    Columns in `item_index` are the axis fields; they get a sharedItems list
    and are stored in the records as integer indices. Everything else is
    stored literally, which is what Excel does for pure value/unused columns.
    """
    n_rows, n_cols = len(rows), len(headers)

    cache_fields = []
    for c in range(n_cols):
        col = [r[c] for r in rows]
        has_blank = any(v is None for v in col)

        if c in item_index:
            uniq = item_index[c][0]
            items = [Missing() if v is None else Text(v=str(v)) for v in uniq]
            si = SharedItems(
                _fields=items,
                containsBlank=True if has_blank else None,
                containsSemiMixedTypes=None,
                count=len(items),
            )
        else:
            nums = [v for v in col if isinstance(v, (int, float)) and not isinstance(v, bool)]
            dates = [v for v in col if isinstance(v, _dt.datetime)]
            strs = [v for v in col if isinstance(v, str)]
            if dates and not strs and not nums:
                si = SharedItems(
                    containsSemiMixedTypes=False,
                    containsNonDate=False,
                    containsDate=True,
                    containsString=False,
                    containsBlank=True if has_blank else None,
                    minDate=min(dates),
                    maxDate=max(dates),
                )
            elif nums and not strs and not dates:
                allint = all(float(v).is_integer() for v in nums)
                si = SharedItems(
                    containsSemiMixedTypes=False,
                    containsString=False,
                    containsNumber=True,
                    containsInteger=True if allint else None,
                    containsBlank=True if has_blank else None,
                    minValue=float(min(nums)),
                    maxValue=float(max(nums)),
                )
            else:
                si = SharedItems(
                    containsBlank=True if has_blank else None,
                    longText=True if any(len(s) > 255 for s in strs) else None,
                )
        cache_fields.append(CacheField(name=headers[c], sharedItems=si))

    # ---- records -------------------------------------------------------
    records = []
    for r in rows:
        fields = []
        for c in range(n_cols):
            v = r[c]
            if c in item_index:
                fields.append(Index(v=item_index[c][1][item_key(v)]))
            elif v is None:
                fields.append(Missing())
            elif isinstance(v, bool):
                fields.append(Boolean(v=v))
            elif isinstance(v, _dt.datetime):
                fields.append(DateTimeField(v=v))
            elif isinstance(v, (int, float)):
                fields.append(Number(v=float(v)))
            else:
                fields.append(Text(v=str(v)))
        records.append(Record(_fields=fields))

    ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    cache = CacheDefinition(
        # Excel re-reads the source on open, so the pivot self-heals
        refreshOnLoad=True,
        saveData=True,
        enableRefresh=True,
        refreshedBy="build_pivot_sheet.py",
        createdVersion=8,
        refreshedVersion=8,
        minRefreshableVersion=3,
        recordCount=n_rows,
        cacheSource=CacheSource(
            type="worksheet",
            worksheetSource=WorksheetSource(ref=ref, sheet=source_sheet),
        ),
        cacheFields=cache_fields,
    )
    cache.records = RecordList(r=records)
    return cache


# --------------------------------------------------------------------------
# Pivot table definition
# --------------------------------------------------------------------------

def build_table(headers, item_index, grid, n_rows_out, num_fmt_id=None,
                name="PivotTable1"):
    n_cols = len(headers)
    row_idx = [headers.index(f) for f in ROW_FIELDS]
    val_idx = [headers.index(c) for c, _, _ in VALUE_FIELDS]

    pivot_fields = []
    for c in range(n_cols):
        if c in row_idx:
            items = [FieldItem(x=i) for i in range(len(item_index[c][0]))]
            items.append(FieldItem(t="default"))          # the subtotal item
            pivot_fields.append(
                PivotField(
                    axis="axisRow",
                    showAll=False,
                    compact=True,
                    outline=True,
                    subtotalTop=True,
                    items=items,
                )
            )
        elif c in val_idx:
            pivot_fields.append(PivotField(dataField=True, showAll=False,
                                           compact=True, outline=True))
        else:
            pivot_fields.append(PivotField(showAll=False, compact=True, outline=True))

    # ---- row items: one <i> per rendered row, minus the header ----------
    # The item index travels with each grid row, so no lookup can go wrong.
    row_items = []
    for level, _label, _vals, _sub, is_grand, idx in grid:
        if is_grand:
            row_items.append(RowColItem(t="grand", x=[Index()]))
        elif level == 0:
            row_items.append(RowColItem(x=[Index(v=idx)]))
        else:
            row_items.append(RowColItem(r=1, x=[Index(v=idx)]))

    # ---- data fields sit on the column axis -----------------------------
    col_items = [RowColItem(i=k, x=[Index(v=k)])
                 for k in range(len(VALUE_FIELDS))]
    data_fields = [
        DataField(name=caption, fld=headers.index(col), subtotal=how,
                  baseField=0, baseItem=0, numFmtId=num_fmt_id)
        for col, how, caption in VALUE_FIELDS
    ]

    last_col = get_column_letter(PIVOT_ANCHOR_COL + len(VALUE_FIELDS))
    ref = (f"{get_column_letter(PIVOT_ANCHOR_COL)}{PIVOT_ANCHOR_ROW}:"
           f"{last_col}{PIVOT_ANCHOR_ROW + n_rows_out}")

    return TableDefinition(
        name=name,
        cacheId=1,
        applyNumberFormats=False,
        applyBorderFormats=False,
        applyFontFormats=False,
        applyPatternFormats=False,
        applyAlignmentFormats=False,
        applyWidthHeightFormats=True,
        dataCaption="Values",
        updatedVersion=8,
        minRefreshableVersion=3,
        createdVersion=8,
        indent=0,
        # Compact form: one "Row Labels" column with indented children, which
        # is what the original Sheet1 uses. These must agree with the
        # per-field compact=True below -- mixing them makes Excel treat the
        # layout as malformed and it drops the PivotTable on open.
        outline=True,
        outlineData=True,
        compact=True,
        compactData=True,
        multipleFieldFilters=False,
        useAutoFormatting=True,
        itemPrintTitles=True,
        location=Location(ref=ref, firstHeaderRow=1, firstDataRow=1, firstDataCol=1),
        pivotFields=pivot_fields,
        rowFields=[RowColField(x=i) for i in row_idx],
        rowItems=row_items,
        colFields=[RowColField(x=-2)],
        colItems=col_items,
        dataFields=data_fields,
        pivotTableStyleInfo=PivotTableStyle(
            name=PIVOT_STYLE,
            showRowHeaders=True,
            showColHeaders=True,
            showRowStripes=False,
            showColStripes=False,
            showLastColumn=True,
        ),
    )


# --------------------------------------------------------------------------
# Rendering the visible grid
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN_TOP = Border(top=Side(style="thin", color="4472C4"))
THIN_BOTH = Border(top=Side(style="thin", color="4472C4"),
                   bottom=Side(style="thin", color="4472C4"))


def render(ws, grid):
    r0, c0 = PIVOT_ANCHOR_ROW, PIVOT_ANCHOR_COL
    captions = ["Row Labels"] + [c for _, _, c in VALUE_FIELDS]

    for j, cap in enumerate(captions):
        cell = ws.cell(row=r0, column=c0 + j, value=cap)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTH
        cell.alignment = Alignment(horizontal="left" if j == 0 else "right")

    for i, (level, label, vals, is_sub, is_grand, _idx) in enumerate(grid, start=1):
        row = r0 + i
        last = is_grand
        text = BLANK_CAPTION if label is None else str(label)
        lc = ws.cell(row=row, column=c0, value=text)
        lc.alignment = Alignment(indent=level)
        if is_sub:
            lc.font = Font(bold=True)
        for j, v in enumerate(vals, start=1):
            vc = ws.cell(row=row, column=c0 + j, value=v)
            vc.number_format = COMMA_0
            if is_sub:
                vc.font = Font(bold=True)
        if is_sub and not last:
            for j in range(len(vals) + 1):
                ws.cell(row=row, column=c0 + j).border = THIN_TOP
        if last:
            for j in range(len(vals) + 1):
                c = ws.cell(row=row, column=c0 + j)
                c.fill = TOTAL_FILL
                c.border = THIN_BOTH

    ws.column_dimensions[get_column_letter(c0)].width = 24
    for j in range(1, len(VALUE_FIELDS) + 1):
        ws.column_dimensions[get_column_letter(c0 + j)].width = max(
            12, len(captions[j]) + 2
        )


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def build(input_path: str,
          output_path: str | None = None,
          source_sheet: str | None = None,
          pivot_sheet: str = "Sheet1",
          quiet: bool = False) -> str:
    """Build the pivot workbook. Importable entry point; returns the output path.

    `input_path` may be the exporter's .csv or an .xlsx.
    """
    out_path = output_path or os.path.splitext(input_path)[0] + "_pivot.xlsx"

    src_name, headers, rows = read_source(input_path, source_sheet)

    missing = [f for f in ROW_FIELDS + [c for c, _, _ in VALUE_FIELDS]
               if f not in headers]
    if missing:
        raise SystemExit(f"columns not found in {src_name!r}: {missing}\n"
                         f"available: {headers}")

    # An all-empty row field yields a pivot that is nothing but "(blank)".
    # Usually means an xlsx formula column with no cached result.
    for f in ROW_FIELDS:
        col = headers.index(f)
        if all(r[col] is None for r in rows):
            print(f"warning: every value in {f!r} is empty. If it is a formula "
                  f"column, open and re-save the workbook in Excel so the "
                  f"cached results are stored, then run this again.",
                  file=sys.stderr)

    indexed = {headers.index(f) for f in ROW_FIELDS}
    item_index = {c: build_item_index(rows, c) for c in indexed}
    tree, grid = aggregate(headers, rows, item_index)

    wb = Workbook()
    wb.remove(wb.active)

    # 1. data sheet (values only -- formulas are resolved)
    data_ws = wb.create_sheet(_safe_sheet_name(src_name))
    data_ws.append(headers)
    for r in rows:
        data_ws.append(r)
    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    )

    # 2. pivot sheet
    pivot_ws = wb.create_sheet(pivot_sheet)
    render(pivot_ws, grid)

    # Reuse the number format we just registered on the cells so a Refresh
    # in Excel keeps the comma formatting instead of reverting to General.
    try:
        num_fmt_id = 164 + wb._number_formats.index(COMMA_0)
    except (ValueError, AttributeError):
        num_fmt_id = None

    cache = build_cache(headers, rows, data_ws.title, item_index)
    table = build_table(headers, item_index, grid, len(grid),
                        num_fmt_id=num_fmt_id)
    table.cache = cache
    pivot_ws.add_pivot(table)

    wb.move_sheet(pivot_sheet, offset=-1)   # pivot first, data behind it
    wb.save(out_path)

    if not quiet:
        print(f"source       : {src_name}  ({len(rows)} rows x {len(headers)} cols)")
        print(f"pivot sheet  : {pivot_sheet}  "
              f"({len(grid)} body rows, anchored at "
              f"{get_column_letter(PIVOT_ANCHOR_COL)}{PIVOT_ANCHOR_ROW})")
        gl, gt = grid[-1][1], grid[-1][2]
        print(f"{gl:<12} : " + "  ".join(
            f"{cap} = {0 if v is None else v:,.0f}"
            for (_, _, cap), v in zip(VALUE_FIELDS, gt)))
        print(f"written      : {out_path}")
    return out_path


def main(argv: Sequence[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="source .csv or .xlsx")
    ap.add_argument("-o", "--output", default=None, help="output .xlsx")
    ap.add_argument("--source-sheet", default=None,
                    help='xlsx: sheet to read (default: first starting '
                         '"social_feed"). csv: name to give the data sheet.')
    ap.add_argument("--pivot-sheet", default="Sheet1")
    args = ap.parse_args(argv)

    build(args.input, args.output, args.source_sheet, args.pivot_sheet)
    return 0


if __name__ == "__main__":
    sys.exit(main())
